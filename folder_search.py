import os
import json
import datetime
import uuid
import subprocess
import requests
from pathlib import Path
import numpy as np

# Document extraction libraries
try:
    import pypdf
except ImportError:
    pypdf = None

try:
    from docx import Document
except ImportError:
    Document = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

SUPPORTED_EXTENSIONS = {'.pdf', '.docx', '.txt', '.xlsx', '.md'}
OLLAMA_GENERATIVE_MODEL = 'llama3:latest'  # For text generation
OLLAMA_EMBEDDING_MODEL = 'mxbai-embed-large:latest'  # For embeddings
OLLAMA_URL = 'http://localhost:11434'

# Key extraction queries for different document types
EXTRACTION_QUERIES = {
    'pricing': ['price', 'cost', 'charge', 'amount', 'fee', 'rate', 'total', 'currency'],
    'parties': ['company', 'client', 'provider', 'vendor', 'supplier', 'organization', 'name', 'contact'],
    'dates': ['date', 'valid', 'expiration', 'deadline', 'delivery', 'arrival', 'departure'],
    'scope': ['service', 'description', 'item', 'content', 'work', 'shipment', 'cargo'],
    'terms': ['payment', 'condition', 'term', 'policy', 'requirement', 'agreement', 'liability']
}

def extract_text_from_file(file_path):
    """Extract text content from various document types."""
    _, ext = os.path.splitext(file_path)
    ext = ext.lower()
    
    try:
        if ext == '.txt' or ext == '.md':
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()
        
        elif ext == '.pdf' and pypdf:
            text = ""
            with open(file_path, 'rb') as f:
                reader = pypdf.PdfReader(f)
                for page in reader.pages:
                    text += page.extract_text()
            return text
        
        elif ext == '.docx' and Document:
            doc = Document(file_path)
            return '\n'.join([para.text for para in doc.paragraphs])
        
        elif ext == '.xlsx' and openpyxl:
            workbook = openpyxl.load_workbook(file_path)
            text = ""
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                text += f"\n--- Sheet: {sheet} ---\n"
                for row in ws.iter_rows(values_only=True):
                    text += '\t'.join([str(cell) if cell is not None else "" for cell in row]) + "\n"
            return text
        
        else:
            return None
    except Exception as e:
        print(f"Error extracting text from {file_path}: {e}")
        return None

def get_embedding(text):
    """Get embedding vector for text using Ollama."""
    if not text or len(text.strip()) < 10:
        return None
    
    try:
        response = requests.post(
            f'{OLLAMA_URL}/api/embed',
            json={'model': OLLAMA_EMBEDDING_MODEL, 'input': text},
            timeout=60
        )
        response.raise_for_status()
        result = response.json()
        embeddings = result.get('embeddings', [])
        if embeddings:
            return embeddings[0]
        return None
    except Exception as e:
        print(f"Error generating embedding: {e}")
        return None

def cosine_similarity(vec1, vec2):
    """Calculate cosine similarity between two vectors."""
    if not vec1 or not vec2:
        return 0.0
    vec1 = np.array(vec1)
    vec2 = np.array(vec2)
    norm1 = np.linalg.norm(vec1)
    norm2 = np.linalg.norm(vec2)
    if norm1 == 0 or norm2 == 0:
        return 0.0
    return float(np.dot(vec1, vec2) / (norm1 * norm2))

def find_key_sections(text, num_sections=3):
    """Find key sections using embeddings by splitting text into chunks."""
    if not text or len(text.strip()) < 200:
        return [text[:2000]] if len(text) > 0 else []
    
    # Split text into sentences
    sentences = [s.strip() for s in text.replace('\n', ' ').split('.') if s.strip()]
    if not sentences:
        return [text[:2000]]
    
    # Create chunks of ~3-5 sentences each
    chunks = []
    chunk_size = 4
    for i in range(0, len(sentences), chunk_size):
        chunk = '. '.join(sentences[i:i+chunk_size])
        if chunk:
            chunks.append(chunk)
    
    if not chunks:
        return [text[:2000]]
    
    # Score chunks and return top ones
    scored_chunks = []
    for chunk in chunks:
        embedding = get_embedding(chunk)
        if embedding:
            scored_chunks.append((chunk, len(chunk)))
    
    # Sort by length and return top sections
    scored_chunks.sort(key=lambda x: x[1], reverse=True)
    return [c[0] for c in scored_chunks[:num_sections]]

def extract_fields_with_embeddings(text, file_name):
    """Extract specific fields using semantic embeddings."""
    extracted_fields = {}
    
    if not text or len(text.strip()) < 100:
        return extracted_fields
    
    # Get text embedding
    text_embedding = get_embedding(text[:1000])
    
    for field, keywords in EXTRACTION_QUERIES.items():
        best_match = None
        best_score = 0.0
        
        # Find chunks matching field keywords
        sentences = [s.strip() for s in text.replace('\n', ' ').split('.') if s.strip()]
        for sentence in sentences:
            # Quick keyword check first
            if any(kw.lower() in sentence.lower() for kw in keywords):
                keyword_embedding = get_embedding(sentence)
                if keyword_embedding:
                    score = cosine_similarity(text_embedding, keyword_embedding)
                    if score > best_score:
                        best_score = score
                        best_match = sentence
        
        if best_match and best_score > 0.3:
            extracted_fields[field] = {
                'value': best_match[:200],
                'confidence': round(best_score, 2)
            }
    
    return extracted_fields

def find_similar_documents(file_path, all_embeddings):
    """Find similar documents based on embeddings."""
    text = extract_text_from_file(file_path)
    if not text:
        return []
    
    current_embedding = get_embedding(text[:1000])
    if not current_embedding:
        return []
    
    similarities = []
    for other_path, other_embedding in all_embeddings.items():
        if other_path == file_path:
            continue
        
        similarity = cosine_similarity(current_embedding, other_embedding)
        if similarity > 0.7:  # High similarity threshold
            similarities.append({
                'path': other_path,
                'name': os.path.basename(other_path),
                'similarity': round(similarity, 2)
            })
    
    return sorted(similarities, key=lambda x: x['similarity'], reverse=True)

def generate_summary_with_ollama(text, file_name):
    """Generate detailed AI summary using local Ollama."""
    if not text or len(text.strip()) < 100:
        return None
    
    # Find key sections using embeddings for faster processing
    key_sections = find_key_sections(text, num_sections=2)
    text_to_summarize = ' '.join(key_sections)[:2000]
    
    prompt = f"""Summarize this document ({file_name}) briefly and provide key details in a structured JSON format.

Document content:
{text}

Return JSON with these sections:
- title: document title/name
- type: document type (freight, report, contract, etc)
- key_info: 2-3 most important details
- parties: who is involved (if applicable)
- purpose: main purpose/objective
- next_steps: any actions needed"""

    try:
        # Try the chat endpoint first (more reliable)
        response = requests.post(
            f'{OLLAMA_URL}/api/chat',
            json={
                'model': OLLAMA_GENERATIVE_MODEL,
                'messages': [
                    {'role': 'user', 'content': prompt}
                ],
                'stream': False
            },
            timeout=300
        )
        
        # If chat endpoint fails, try generate endpoint
        if response.status_code != 200:
            response = requests.post(
                f'{OLLAMA_URL}/api/generate',
                json={'model': OLLAMA_GENERATIVE_MODEL, 'prompt': prompt, 'stream': False},
                timeout=300
            )
        
        response.raise_for_status()
        result = response.json()
        
        # Extract response based on endpoint
        if 'message' in result:
            summary_text = result['message']['content']
        else:
            summary_text = result.get('response', '')
        
        # Try to parse as JSON, otherwise return as text
        try:
            parsed = json.loads(summary_text)
            return parsed
        except json.JSONDecodeError:
            # Try to extract JSON from markdown code blocks
            if '```json' in summary_text:
                json_str = summary_text.split('```json')[1].split('```')[0].strip()
                try:
                    return json.loads(json_str)
                except json.JSONDecodeError:
                    pass
            elif '```' in summary_text:
                json_str = summary_text.split('```')[1].split('```')[0].strip()
                try:
                    return json.loads(json_str)
                except json.JSONDecodeError:
                    pass
            
            # If no JSON found, return as raw summary
            return {"raw_summary": summary_text}
    
    except requests.exceptions.ConnectionError:
        print(f"Warning: Could not connect to Ollama at {OLLAMA_URL}. Make sure Ollama is running.")
        return None
    except Exception as e:
        print(f"Error generating summary with Ollama: {e}")
        return None

def get_file_info(file_path, generate_ai_summary=True, use_embeddings=True, all_embeddings=None):
    """Get detailed file information including AI summary and extracted fields."""
    try:
        stat_info = os.stat(file_path)
        file_size_bytes = stat_info.st_size
        file_size_human = format_file_size(file_size_bytes)
        
        _, ext = os.path.splitext(file_path)
        
        file_info = {
            "name": os.path.basename(file_path),
            "path": file_path,
            "extension": ext.lower(),
            "size_bytes": file_size_bytes,
            "size_human": file_size_human
        }
        
        # Generate AI summary if enabled and file type is supported
        if generate_ai_summary and ext.lower() in SUPPORTED_EXTENSIONS:
            text_content = extract_text_from_file(file_path)
            if text_content:
                # Extract fields using embeddings
                if use_embeddings:
                    extracted_fields = extract_fields_with_embeddings(text_content, os.path.basename(file_path))
                    if extracted_fields:
                        file_info["extracted_fields"] = extracted_fields
                
                # Generate summary
                summary = generate_summary_with_ollama(text_content, os.path.basename(file_path))
                if summary:
                    file_info["ai_summary"] = summary
                
                # Find similar documents
                if use_embeddings and all_embeddings is not None:
                    similar = find_similar_documents(file_path, all_embeddings)
                    if similar:
                        file_info["similar_documents"] = similar
                
                # Store embedding for similarity calculations
                if all_embeddings is not None:
                    embedding = get_embedding(text_content[:1000])
                    if embedding:
                        all_embeddings[file_path] = embedding
        
        return file_info
    
    except Exception as e:
        print(f"Error getting file info for {file_path}: {e}")
        return None

def format_file_size(bytes_size):
    """Convert bytes to human-readable format."""
    for unit in ['B', 'KB', 'MB', 'GB']:
        if bytes_size < 1024.0:
            return f"{bytes_size:.2f} {unit}"
        bytes_size /= 1024.0
    return f"{bytes_size:.2f} TB"

def get_folder_structure(root_dir, max_depth=6, generate_summaries=True, use_embeddings=True):
    all_embeddings = {} if use_embeddings else None
    
    def _walk(dir_path, depth):
        if depth > max_depth:
            return None
        structure = {}
        files_list = []
        try:
            entries = os.listdir(dir_path)
        except PermissionError:
            return None
        for entry in entries:
            full_path = os.path.join(dir_path, entry)
            if os.path.isdir(full_path):
                sub = _walk(full_path, depth + 1)
                if sub is not None:
                    structure[entry] = sub
            else:
                file_info = get_file_info(full_path, generate_ai_summary=generate_summaries, 
                                         use_embeddings=use_embeddings, all_embeddings=all_embeddings)
                if file_info:
                    files_list.append(file_info)
        if files_list:
            structure['file_list'] = files_list
        return structure
    
    structure = _walk(root_dir, 0)
    return structure, all_embeddings

if __name__ == "__main__":
    start_folder = input("Enter the starting folder path: ")
    if not os.path.isdir(start_folder):
        print("Invalid directory")
        exit()
    
    generate_ai = input("Generate AI summaries? (y/n, default: n): ").lower().strip() == 'y'
    use_embeddings = input("Use embeddings for field extraction & similarity? (y/n, default: n): ").lower().strip() == 'y' if generate_ai else False
    
    if generate_ai:
        print("Checking Ollama connection...")
        try:
            response = requests.get(f'{OLLAMA_URL}/api/tags', timeout=5)
            if response.status_code == 200:
                models_data = response.json().get('models', [])
                if models_data:
                    print(f"✓ Ollama is running with {len(models_data)} model(s)")
                    print("Available models:")
                    for model in models_data:
                        print(f"  - {model.get('name', 'unknown')}")
                    
                    # Check if models exist
                    available_names = [m.get('name', '') for m in models_data]
                    if not any(OLLAMA_GENERATIVE_MODEL in name for name in available_names):
                        print(f"\n⚠ Model '{OLLAMA_GENERATIVE_MODEL}' not found. Using first available model.")
                        OLLAMA_GENERATIVE_MODEL = models_data[0].get('name', 'llama3:latest')
                    
                    if use_embeddings and not any(OLLAMA_EMBEDDING_MODEL in name for name in available_names):
                        print(f"⚠ Embedding model '{OLLAMA_EMBEDDING_MODEL}' not found.")
                        use_embeddings = False
                    
                    print(f"Using generative model: {OLLAMA_GENERATIVE_MODEL}")
                    if use_embeddings:
                        print(f"Using embedding model: {OLLAMA_EMBEDDING_MODEL}")
                else:
                    print(f"⚠ Ollama is running but no models found. Please pull a model first.")
                    print(f"Run: ollama pull llama3")
                    generate_ai = False
            else:
                print("⚠ Ollama connection failed")
                generate_ai = False
        except Exception as e:
            print(f"⚠ Could not connect to Ollama: {e}")
            print(f"Make sure Ollama is running at {OLLAMA_URL}")
            print(f"Start Ollama with: ollama serve")
            generate_ai = False
    
    print("\nScanning folder structure...")
    structure, all_embeddings = get_folder_structure(start_folder, 4, generate_summaries=generate_ai, use_embeddings=use_embeddings)
    
    folder_name = os.path.basename(start_folder)
    output_file = os.path.join(start_folder, f"{folder_name}-structure.json")
    
    last_modified = None
    if os.path.exists(output_file):
        last_modified = datetime.datetime.fromtimestamp(os.path.getmtime(output_file)).isoformat()
    
    generated_at = datetime.datetime.now().isoformat()
    short_guid = str(uuid.uuid4())[:8]
    
    structure['generated_at'] = generated_at
    structure['guid'] = short_guid
    structure['ai_summaries_enabled'] = generate_ai
    structure['embeddings_enabled'] = use_embeddings
    structure['ollama_generative_model'] = OLLAMA_GENERATIVE_MODEL if generate_ai else None
    structure['ollama_embedding_model'] = OLLAMA_EMBEDDING_MODEL if use_embeddings else None
    
    if last_modified:
        structure['last_modified'] = last_modified
    
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(structure, f, indent=2, ensure_ascii=False)
    
    print(f"✓ JSON output saved to {output_file}")